# -*- coding: utf-8 -*-
"""Generadores de archivos XLSX para reportes descargables."""

import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Verde militar para encabezados
HEADER_BG = "4A5D23"
HEADER_FG = "FFFFFF"


def _style_header_row(ws, row: int, num_cols: int) -> None:
    """Aplica estilo de encabezado a una fila."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    """Ajusta el ancho de columnas automáticamente."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 4, 50
        )


def generate_records_xlsx(
    rows: list[dict[str, Any]],
    filters_applied: dict[str, Any],
) -> bytes:
    """Genera XLSX de registros de arqueo con filtros aplicados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros de Arqueo"

    # Metadatos de filtros
    ws["A1"] = "Reporte: Registros de Arqueo"
    ws["A1"].font = Font(bold=True, size=12)
    filter_str = " | ".join(
        f"{k}: {v}" for k, v in filters_applied.items() if v is not None
    )
    ws["A2"] = f"Filtros: {filter_str or 'ninguno'}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.append([])

    headers = [
        "UID", "Tipo Registro", "UID Origen",
        "Fecha Arqueo", "Empresa", "Bóveda",
        "Comprobante", "Referencia", "Sucursal",
        "Tipo Movimiento", "Entradas", "Salidas",
        # Billetes
        "$1,000", "$500", "$200", "$100", "$50", "$20",
        # Monedas
        "$10", "$5", "$2", "$1", "$0.50", "$0.20", "$0.10",
        "Fecha Registro", "Estado",
    ]
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    # Etiqueta legible para la columna de trazabilidad
    def _record_kind(r: dict[str, Any]) -> str:
        if not r.get("is_counterpart"):
            return "Original"
        ctype = r.get("counterpart_type")
        if ctype == "cancellation":
            return "Cancelación"
        if ctype == "modification":
            return "Modificación"
        return "Contrapartida"

    # Sombreado gris claro para filas de contrapartida (trazabilidad de
    # cambios). Hace evidente al lector qué se canceló o se sustituyó.
    counterpart_fill = PatternFill("solid", fgColor="F0E6E6")

    for row in rows:
        excel_row_idx = ws.max_row + 1
        ws.append([
            row.get("record_uid", ""),
            _record_kind(row),
            row.get("original_record_uid", "") or "",
            row.get("arqueo_date", ""),
            row.get("company_name", ""),
            row.get("vault_code", ""),
            row.get("voucher", ""),
            row.get("reference", ""),
            row.get("branch_name", ""),
            row.get("movement_type_name", ""),
            row.get("entries", 0),
            row.get("withdrawals", 0),
            # Billetes
            row.get("bill_1000", 0),
            row.get("bill_500", 0),
            row.get("bill_200", 0),
            row.get("bill_100", 0),
            row.get("bill_50", 0),
            row.get("bill_20", 0),
            # Monedas
            row.get("coin_100", 0),
            row.get("coin_50", 0),
            row.get("coin_20", 0),
            row.get("coin_10", 0),
            row.get("coin_5", 0),
            row.get("coin_2", 0),
            row.get("coin_1", 0),
            row.get("coin_050", 0),
            row.get("coin_020", 0),
            row.get("coin_010", 0),
            row.get("record_date", ""),
            row.get("header_status", ""),
        ])
        if row.get("is_counterpart"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=excel_row_idx, column=c).fill = counterpart_fill

    # Formato numérico para Entradas/Salidas y las 16 denominaciones
    # (columnas K..AB → 11 a 28 en 1-indexed).
    for row_cells in ws.iter_rows(min_row=5, min_col=11, max_col=28):
        for cell in row_cells:
            cell.number_format = '#,##0.00'

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
